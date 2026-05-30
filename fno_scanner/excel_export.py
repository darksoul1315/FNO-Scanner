import os
import pandas as pd
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

if HAS_OPENPYXL:
    HEADER_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    HEADER_FILL = PatternFill(start_color='1B2A4A', end_color='1B2A4A', fill_type='solid')
    HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
    DATA_FONT = Font(name='Calibri', size=10)
    DATA_CENTER = Alignment(horizontal='center', vertical='center')
    DATA_LEFT = Alignment(horizontal='left', vertical='center')
    THIN_BORDER = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9'),
    )
    FILL_BULLISH = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    FILL_BEARISH = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    FILL_NEUTRAL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    FILL_ALT_ROW = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

    SCORE_FILLS = {
        'high': PatternFill(start_color='00B050', end_color='00B050', fill_type='solid'),
        'mid': PatternFill(start_color='92D050', end_color='92D050', fill_type='solid'),
        'dev': PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid'),
        'low': PatternFill(start_color='FF6347', end_color='FF6347', fill_type='solid'),
    }

    OI_FILLS = {
        'Long Buildup': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
        'Short Covering': PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid'),
        'Short Buildup': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
        'Long Unwinding': PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid'),
    }

    PHASE_FILLS = {
        'Leading': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
        'Improving': PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid'),
        'Weakening': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
        'Lagging': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
    }

    FILL_FAIL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    FILL_PASS_ROW = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

    STABILITY_FILLS = {
        'HIGH': PatternFill(start_color='00B050', end_color='00B050', fill_type='solid'),
        'MED': PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid'),
        'LOW': PatternFill(start_color='FF6347', end_color='FF6347', fill_type='solid'),
    }

    TITLE_FONT = Font(name='Calibri', bold=True, size=14, color='1B2A4A')
    SECTION_FONT = Font(name='Calibri', bold=True, size=12, color='1B2A4A')
    SUBTITLE_FONT = Font(name='Calibri', size=10, italic=True, color='666666')


def _write_title(ws, row, text, cols=10):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 30


def _write_subtitle(ws, row, text, cols=10):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = SUBTITLE_FONT
    cell.alignment = Alignment(horizontal='center')


def _write_headers(ws, row, headers):
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 25


def _write_data_row(ws, row, values, is_alt=False):
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = DATA_FONT
        cell.alignment = DATA_CENTER
        cell.border = THIN_BORDER
        if is_alt:
            cell.fill = FILL_ALT_ROW


def _auto_width(ws, col_count, start_row=1, end_row=None):
    if end_row is None:
        end_row = ws.max_row
    for col_idx in range(1, col_count + 1):
        max_len = 8
        for r in range(start_row, end_row + 1):
            v = ws.cell(row=r, column=col_idx).value
            if v is not None:
                max_len = max(max_len, len(str(v)) + 2)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len, 30)


def _apply_score_fill(cell):
    try:
        v = int(cell.value)
        if v >= 55:
            cell.fill = SCORE_FILLS['high']
            cell.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
        elif v >= 40:
            cell.fill = SCORE_FILLS['mid']
        elif v >= 25:
            cell.fill = SCORE_FILLS['dev']
        else:
            cell.fill = SCORE_FILLS['low']
            cell.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    except (ValueError, TypeError):
        pass


def _apply_bias_fill(cell):
    v = str(cell.value)
    if v == 'Bullish':
        cell.fill = FILL_BULLISH
        cell.font = Font(name='Calibri', size=10, bold=True, color='006100')
    elif v == 'Bearish':
        cell.fill = FILL_BEARISH
        cell.font = Font(name='Calibri', size=10, bold=True, color='9C0006')
    elif v == 'Neutral':
        cell.fill = FILL_NEUTRAL
        cell.font = Font(name='Calibri', size=10, bold=True, color='9C6500')


def _build_full_results_sheet(ws, df, market_regime, stats):
    rename_map = {
        '_sub_liq': 'Liq(10)', '_sub_oi': 'OI(15)', '_sub_mom': 'Mom(15)',
        '_sub_rs': 'RS(10)', '_sub_vol': 'Vol(15)', '_sub_vola': 'Vola(10)',
        '_sub_sm': 'SM(15)', '_sub_opt': 'Opt(10)',
        '_compressed': 'Compressed?', '_pocket_pivot': 'PocketPivot?',
        '_accumulation': 'Accum?', '_nr7': 'NR7?',
        '_above_vwap': 'AboveVWAP?', '_trend': 'Trend',
    }
    export_df = df.rename(columns=rename_map).copy()

    keep_cols = [
        '#', 'Symbol', 'Sector', 'CMP', 'Chg%', 'VolRatio',
        'Del%', 'FUT_OI', 'FUT_CONTRACTS', 'ATR_Exp%',
        'OI_Chg%', 'OI_Class', 'PCR', 'IV%ile', 'RS_Rank',
        'Score', 'Stability', 'Setup', 'Liq_Zone', 'Bias', 'ML_Conf', 'Data_Source', 'Liq(10)',
        'OI(15)', 'Mom(15)', 'RS(10)', 'Vol(15)', 'Vola(10)',
        'SM(15)', 'Opt(10)', 'Compressed?', 'PocketPivot?',
        'Accum?', 'NR7?', 'AboveVWAP?', 'Trend'
    ]

    all_cols = [c for c in keep_cols if c in export_df.columns]
    export_df = export_df[all_cols]
    ncols = len(all_cols)

    _write_title(ws, 1, f"INSTITUTIONAL F&O SCANNER — RESULTS — {datetime.now().strftime('%d %b %Y %H:%M')}", ncols)
    regime_text = market_regime.get('regime', 'neutral').upper()
    _write_subtitle(ws, 2, f"Market Regime: {regime_text} | Analyzed: {stats.get('analyzed', 0)} | Results: {len(export_df)}", ncols)

    hdr_row = 4
    _write_headers(ws, hdr_row, all_cols)
    ws.freeze_panes = f'A{hdr_row + 1}'

    col_map = {name: idx + 1 for idx, name in enumerate(all_cols)}

    for r_idx, (_, row_data) in enumerate(export_df.iterrows(), start=hdr_row + 1):
        is_alt = (r_idx - hdr_row) % 2 == 0
        for c_idx, col in enumerate(all_cols, 1):
            val = row_data[col]
            if isinstance(val, bool):
                val = 'YES' if val else ''
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = DATA_FONT
            cell.alignment = DATA_CENTER
            cell.border = THIN_BORDER
            if is_alt:
                cell.fill = FILL_ALT_ROW

        if 'Score' in col_map:
            _apply_score_fill(ws.cell(row=r_idx, column=col_map['Score']))
        if 'Bias' in col_map:
            _apply_bias_fill(ws.cell(row=r_idx, column=col_map['Bias']))
        if 'OI_Class' in col_map:
            oi_cell = ws.cell(row=r_idx, column=col_map['OI_Class'])
            if str(oi_cell.value) in OI_FILLS:
                oi_cell.fill = OI_FILLS[str(oi_cell.value)]
        if 'Chg%' in col_map:
            chg_cell = ws.cell(row=r_idx, column=col_map['Chg%'])
            try:
                v = float(chg_cell.value)
                if v > 0:
                    chg_cell.font = Font(name='Calibri', size=10, color='006100')
                elif v < 0:
                    chg_cell.font = Font(name='Calibri', size=10, color='9C0006')
            except (ValueError, TypeError):
                pass

        if 'Stability' in col_map:
            st_cell = ws.cell(row=r_idx, column=col_map['Stability'])
            if str(st_cell.value) in STABILITY_FILLS:
                st_cell.fill = STABILITY_FILLS[str(st_cell.value)]
                st_cell.font = Font(name='Calibri', size=10, bold=True)

    _auto_width(ws, ncols, hdr_row, hdr_row + len(export_df))


def _build_data_collection_sheet(wb, stats):
    ws = wb.create_sheet("Phase 1 — Data Collection")
    _write_title(ws, 1, "PHASE 1: DATA COLLECTION SUMMARY", 4)
    _write_headers(ws, 3, ['Metric', 'Value', 'Status', 'Notes'])

    rows = [
        ['Total F&O Stocks', stats.get('total', 0), 'OK', 'NSE F&O universe'],
        ['Data Fetched OK', stats.get('data_ok', 0), 'OK' if stats.get('data_ok', 0) > 0 else 'WARN', 'Valid OHLCV data'],
        ['Passed Liquidity', stats.get('liquidity_pass', 0), 'OK', 'Min price/volume/value filters'],
        ['Fully Analyzed', stats.get('analyzed', 0), 'OK', 'Complete pipeline pass'],
        ['Errors/Skipped', stats.get('failed', 0), 'WARN' if stats.get('failed', 0) > 5 else 'OK', 'Data/API errors'],
    ]

    for i, row in enumerate(rows, start=4):
        _write_data_row(ws, i, row, i % 2 == 0)
        status_cell = ws.cell(row=i, column=3)
        if str(status_cell.value) == 'OK':
            status_cell.fill = FILL_BULLISH
        else:
            status_cell.fill = FILL_NEUTRAL

    _auto_width(ws, 4)


def _build_sector_rotation_sheet(wb, sector_metrics):
    from .sector_rotation import rank_sectors
    ws = wb.create_sheet("Phase 2 — Sector Rotation")
    _write_title(ws, 1, "PHASE 2: SECTOR ROTATION ANALYSIS", 7)

    headers = ['Sector', 'Phase', 'RS (1M)', 'RS (3M)', 'Return 1M%', 'Breadth%', 'MomScore']
    _write_headers(ws, 3, headers)

    ranked = rank_sectors(sector_metrics)
    for r_idx, (sector, m) in enumerate(ranked, start=4):
        vals = [sector, m['phase'], round(m['rs_1m'], 4), round(m.get('rs_3m', 0), 4),
                round(m.get('return_1m', 0), 2), round(m['breadth_pct'], 1), m['momentum_score']]
        _write_data_row(ws, r_idx, vals, r_idx % 2 == 0)
        phase_cell = ws.cell(row=r_idx, column=2)
        if m['phase'] in PHASE_FILLS:
            phase_cell.fill = PHASE_FILLS[m['phase']]

    _auto_width(ws, 7)


def _build_pipeline_stats_sheet(wb, df, stats):
    ws = wb.create_sheet("Phase 3 — Pipeline Stats")
    _write_title(ws, 1, "PHASE 3: STOCK ANALYSIS PIPELINE", 5)
    _write_headers(ws, 3, ['Metric', 'Count', '%', 'Description', 'Status'])

    total = max(stats.get('total', 1), 1)
    rows = [
        ['Total Universe', stats.get('total', 0), '100%', 'F&O eligible stocks', 'BASE'],
        ['Data OK', stats.get('data_ok', 0), f"{stats.get('data_ok',0)/total*100:.0f}%", 'Valid OHLCV fetched', 'OK'],
        ['Liquidity Pass', stats.get('liquidity_pass', 0), f"{stats.get('liquidity_pass',0)/total*100:.0f}%", 'Price/Vol/Value filter', 'OK'],
        ['Analyzed', stats.get('analyzed', 0), f"{stats.get('analyzed',0)/total*100:.0f}%", 'Full pipeline complete', 'OK'],
        ['Errors', stats.get('failed', 0), f"{stats.get('failed',0)/total*100:.0f}%", 'Pipeline errors', 'WARN' if stats.get('failed',0)>0 else 'OK'],
    ]

    for i, row in enumerate(rows, start=4):
        _write_data_row(ws, i, row, i % 2 == 0)

    r = 10
    ws.cell(row=r, column=1, value="SECTOR DISTRIBUTION OF RESULTS").font = SECTION_FONT
    r += 1
    _write_headers(ws, r, ['Sector', 'Count', 'Avg Score', 'Top Stock', 'Top Score'])
    r += 1

    if not df.empty and 'Sector' in df.columns:
        for sector in df['Sector'].unique():
            sec_df = df[df['Sector'] == sector].sort_values('Score', ascending=False)
            vals = [sector, len(sec_df), round(sec_df['Score'].mean(), 1),
                    sec_df.iloc[0]['Symbol'], sec_df.iloc[0]['Score']]
            _write_data_row(ws, r, vals, r % 2 == 0)
            r += 1

    _auto_width(ws, 5)


def _build_scoring_ranking_sheet(wb, df):
    ws = wb.create_sheet("Phase 4 — Scoring & Ranking")
    cols = ['#', 'Symbol', 'Sector', 'CMP', 'Chg%', 'Score', 'Setup', 'Bias', 'OI_Class', 'RS_Rank']
    avail = [c for c in cols if c in df.columns]

    _write_title(ws, 1, "PHASE 4: TOP INSTITUTIONAL SETUPS", len(avail))
    _write_headers(ws, 3, avail)

    top = df.head(30)
    for r_idx, (_, row_data) in enumerate(top.iterrows(), start=4):
        vals = [row_data[c] for c in avail]
        _write_data_row(ws, r_idx, vals, r_idx % 2 == 0)
        col_map = {name: idx + 1 for idx, name in enumerate(avail)}
        if 'Score' in col_map:
            _apply_score_fill(ws.cell(row=r_idx, column=col_map['Score']))
        if 'Bias' in col_map:
            _apply_bias_fill(ws.cell(row=r_idx, column=col_map['Bias']))
        if 'OI_Class' in col_map:
            oi_cell = ws.cell(row=r_idx, column=col_map['OI_Class'])
            if str(oi_cell.value) in OI_FILLS:
                oi_cell.fill = OI_FILLS[str(oi_cell.value)]

    _auto_width(ws, len(avail))


def _build_category_sheet(wb, df):
    ws = wb.create_sheet("Category Breakdown")
    summary_cols = ['Symbol', 'Sector', 'CMP', 'Chg%', 'Score', 'Setup', 'OI_Class', 'VolRatio', 'RS_Rank']
    avail = [c for c in summary_cols if c in df.columns]
    ncols = max(len(avail), 6)

    _write_title(ws, 1, "CATEGORY BREAKDOWN — ALL CLASSIFICATIONS", ncols)
    row = 3

    categories = []

    bullish = df[df['Bias'] == 'Bullish'] if 'Bias' in df.columns else pd.DataFrame()
    if not bullish.empty:
        categories.append(('BULLISH SETUPS', bullish, FILL_BULLISH))

    bearish = df[df['Bias'] == 'Bearish'] if 'Bias' in df.columns else pd.DataFrame()
    if not bearish.empty:
        categories.append(('BEARISH SETUPS', bearish, FILL_BEARISH))

    squeeze = df[df['Setup'].str.contains('Short Squeeze|Short Covering', na=False, case=False)] if 'Setup' in df.columns else pd.DataFrame()
    if not squeeze.empty:
        categories.append(('SHORT SQUEEZE CANDIDATES', squeeze, PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')))

    compressed = df[df.get('_compressed', pd.Series(dtype=bool)) == True] if '_compressed' in df.columns else pd.DataFrame()
    if not compressed.empty:
        categories.append(('COMPRESSION BREAKOUT READY', compressed, PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')))

    pp = df[df.get('_pocket_pivot', pd.Series(dtype=bool)) == True] if '_pocket_pivot' in df.columns else pd.DataFrame()
    if not pp.empty:
        categories.append(('POCKET PIVOTS — Institutional Buying', pp, PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')))

    accum = df[df.get('_accumulation', pd.Series(dtype=bool)) == True] if '_accumulation' in df.columns else pd.DataFrame()
    if not accum.empty:
        categories.append(('ACCUMULATION DETECTED', accum, PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')))

    nr7 = df[df.get('_nr7', pd.Series(dtype=bool)) == True] if '_nr7' in df.columns else pd.DataFrame()
    if not nr7.empty:
        categories.append(('NR7 — Narrow Range', nr7, PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')))

    for cat_name, cat_df, cat_fill in categories:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        cell = ws.cell(row=row, column=1, value=f"{cat_name} ({len(cat_df)} stocks)")
        cell.font = Font(name='Calibri', bold=True, size=11, color='1B2A4A')
        cell.fill = cat_fill
        cell.alignment = Alignment(horizontal='left', vertical='center')
        row += 1

        _write_headers(ws, row, avail)
        row += 1

        for _, r_data in cat_df.head(15).iterrows():
            vals = [r_data[c] for c in avail]
            _write_data_row(ws, row, vals, row % 2 == 0)
            score_idx = avail.index('Score') + 1 if 'Score' in avail else None
            if score_idx:
                _apply_score_fill(ws.cell(row=row, column=score_idx))
            row += 1

        row += 1

    _auto_width(ws, ncols)


def _build_legend_sheet(wb):
    ws = wb.create_sheet("Scoring Legend")
    _write_title(ws, 1, "SCORING SYSTEM & INTERPRETATION GUIDE", 5)

    row = 3
    _write_headers(ws, row, ['Component', 'Max Points', 'Weight', 'What It Measures', 'Key Signals'])
    row += 1

    legend_data = [
        ['Liquidity', 10, '10%', 'How tradeable is the stock', 'Avg traded value, volume ratio, price level'],
        ['OI Score', 15, '15%', 'Open Interest dynamics', 'OI change, Long/Short buildup classification'],
        ['Momentum', 15, '15%', 'Directional momentum strength', 'RSI, MACD, ROC, breakout proximity'],
        ['Relative Strength', 10, '10%', 'Performance vs NIFTY benchmark', 'Multi-period RS, sector bonus'],
        ['Volume', 15, '15%', 'Institutional volume signatures', 'Vol ratio, accumulation, pocket pivot, OBV'],
        ['Volatility', 10, '10%', 'Compression/expansion cycle', 'ATR ratio, Bollinger BW, NR7, inside day'],
        ['Smart Money', 15, '15%', 'Institutional flow detection', 'Delivery %, absorption, volume dry-up'],
        ['Options', 10, '10%', 'Options positioning signals', 'PCR, IV percentile, gamma zone, max pain'],
    ]

    for vals in legend_data:
        _write_data_row(ws, row, vals, row % 2 == 0)
        row += 1

    row += 2
    ws.cell(row=row, column=1, value="SCORE INTERPRETATION").font = SECTION_FONT
    row += 1
    _write_headers(ws, row, ['Range', 'Label', 'Action', 'Color', 'Description'])
    row += 1

    interp = [
        ['55-100', 'HIGH CONVICTION', 'Strong Trade', 'Green', 'Multiple institutional signals aligned'],
        ['40-54', 'MODERATE', 'Watch / Partial', 'Light Green', 'Developing institutional interest'],
        ['25-39', 'DEVELOPING', 'Watchlist Only', 'Yellow/Orange', 'Early signals, needs confirmation'],
        ['0-24', 'WEAK', 'Avoid', 'Red', 'No clear institutional pattern'],
    ]
    for vals in interp:
        _write_data_row(ws, row, vals, row % 2 == 0)
        row += 1

    row += 2
    ws.cell(row=row, column=1, value="OI CLASSIFICATION").font = SECTION_FONT
    row += 1
    _write_headers(ws, row, ['OI Class', 'Price', 'OI', 'Bias', 'Interpretation'])
    row += 1

    oi_data = [
        ['Long Buildup', 'Rising', 'Rising', 'Bullish', 'New longs entering'],
        ['Short Covering', 'Rising', 'Falling', 'Mildly Bullish', 'Shorts exiting'],
        ['Short Buildup', 'Falling', 'Rising', 'Bearish', 'New shorts entering'],
        ['Long Unwinding', 'Falling', 'Falling', 'Mildly Bearish', 'Longs exiting'],
    ]
    for vals in oi_data:
        _write_data_row(ws, row, vals, row % 2 == 0)
        oi_cell = ws.cell(row=row, column=1)
        if vals[0] in OI_FILLS:
            oi_cell.fill = OI_FILLS[vals[0]]
        row += 1

    _auto_width(ws, 5)


def _build_full_universe_sheet(wb, full_universe_df, stats):
    ws = wb.create_sheet("Full Universe")
    ncols = 14
    _write_title(ws, 1, "FULL UNIVERSE — ALL F&O STOCKS", ncols)
    total = stats.get('total', 0)
    passed = stats.get('liquidity_pass', 0)
    failed = total - passed
    _write_subtitle(ws, 2, f"Total: {total} | Passed Liquidity: {passed} | Failed: {failed}", ncols)

    headers = ['#', 'Symbol', 'Sector', 'CMP', 'Avg_Vol_20d', 'Status', 'Score', 'Stability', 'OI_Chg%', 'OI_Class', 'PCR', 'Del%', 'Bias', 'Data_Source']
    hdr_row = 4
    _write_headers(ws, hdr_row, headers)
    ws.freeze_panes = f'A{hdr_row + 1}'

    for r_idx, (_, row_data) in enumerate(full_universe_df.iterrows(), start=hdr_row + 1):
        is_fail = str(row_data.get('Status', '')).startswith('FAIL')
        vals = [
            r_idx - hdr_row,
            row_data.get('Symbol', ''),
            row_data.get('Sector', ''),
            row_data.get('CMP', ''),
            row_data.get('Avg_Vol_20d', ''),
            row_data.get('Status', ''),
            row_data.get('Score', ''),
            row_data.get('Stability', ''),
            row_data.get('OI_Chg%', ''),
            row_data.get('OI_Class', ''),
            row_data.get('PCR', ''),
            row_data.get('Del%', ''),
            row_data.get('Bias', ''),
            row_data.get('Data_Source', 'PROXY'),
        ]
        for c_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = DATA_FONT
            cell.alignment = DATA_CENTER
            cell.border = THIN_BORDER
            if is_fail:
                cell.fill = FILL_FAIL
        if not is_fail:
            status_cell = ws.cell(row=r_idx, column=6)
            status_cell.fill = FILL_PASS_ROW
            if row_data.get('Score'):
                _apply_score_fill(ws.cell(row=r_idx, column=7))
        stability_val = str(row_data.get('Stability', ''))
        if stability_val in STABILITY_FILLS:
            st_cell = ws.cell(row=r_idx, column=8)
            st_cell.fill = STABILITY_FILLS[stability_val]
            st_cell.font = Font(name='Calibri', size=10, bold=True)

    _auto_width(ws, ncols, hdr_row, hdr_row + len(full_universe_df))


def _build_monte_carlo_sheet(wb, df):
    ws = wb.create_sheet("Monte Carlo Stability")

    cols = ['#', 'Symbol', 'Sector', 'Score', 'Stability', 'MC Mean', 'MC Std', 'MC CV', 'Drop Risk%', 'Upside%']
    avail = [c for c in cols if c in df.columns]
    ncols = len(avail)

    _write_title(ws, 1, "MONTE CARLO STABILITY ANALYSIS — 200 Perturbation Iterations", ncols)
    _write_subtitle(ws, 2, "CV = Std/Mean. HIGH=CV<10% stable | MED=CV 10-20% | LOW=CV>20% unreliable", ncols)

    hdr_row = 4
    _write_headers(ws, hdr_row, avail)
    ws.freeze_panes = f'A{hdr_row + 1}'

    for r_idx, (_, row) in enumerate(df.iterrows(), start=hdr_row + 1):
        vals = [
            r_idx - hdr_row,
            row.get('Symbol', ''),
            row.get('Sector', ''),
            row.get('Score', ''),
            row.get('Stability', ''),
            row.get('_mc_mean', ''),
            row.get('_mc_std', ''),
            row.get('_mc_cv', ''),
            row.get('_mc_drop_risk', ''),
            row.get('_mc_upside', ''),
        ]
        vals = vals[:ncols]
        for c_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = DATA_FONT
            cell.alignment = DATA_CENTER
            cell.border = THIN_BORDER

        st_idx = avail.index('Stability') + 1 if 'Stability' in avail else None
        if st_idx:
            st_cell = ws.cell(row=r_idx, column=st_idx)
            if str(st_cell.value) in STABILITY_FILLS:
                st_cell.fill = STABILITY_FILLS[str(st_cell.value)]
                st_cell.font = Font(name='Calibri', size=10, bold=True)

    _auto_width(ws, ncols, hdr_row, hdr_row + len(df))


def export_comprehensive_excel(scanner, filename=None):
    if not HAS_OPENPYXL:
        print("  [!] openpyxl not installed. Run: pip install openpyxl")
        return None

    df = scanner.results
    if isinstance(df, list) or (isinstance(df, pd.DataFrame) and df.empty):
        print("  No results to export.")
        return None

    if filename is None:
        filename = "fno_scanner_report.xlsx"

    filepath = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', filename)

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "FNO Scanner Results"
    _build_full_results_sheet(ws1, df, scanner.market_regime, scanner.stats)

    _build_data_collection_sheet(wb, scanner.stats)

    if not scanner.full_universe.empty:
        _build_full_universe_sheet(wb, scanner.full_universe, scanner.stats)

    if scanner.sector_metrics:
        _build_sector_rotation_sheet(wb, scanner.sector_metrics)

    _build_monte_carlo_sheet(wb, df)
    _build_pipeline_stats_sheet(wb, df, scanner.stats)
    _build_scoring_ranking_sheet(wb, df)
    _build_category_sheet(wb, df)
    _build_legend_sheet(wb)

    wb.save(filepath)
    abs_path = os.path.abspath(filepath)
    print(f"\n  \u2713 Comprehensive Excel report saved: {filename}")
    print(f"    \U0001f4ca Sheets: {len(wb.sheetnames)}")
    for i, name in enumerate(wb.sheetnames, 1):
        print(f"       {i}. {name}")

    return filepath
